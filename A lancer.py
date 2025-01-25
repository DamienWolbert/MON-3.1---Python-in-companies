import pandas as pd
from openpyxl import load_workbook

path_excel = r"C:\Users\damie\Documents\GitHub\MON-3.1---Python-in-companies\Tests 2.xlsm"
nom_feuille = "Test 2"

#Ouverture de la feuille
fichier = load_workbook(path_excel)
feuille = fichier[nom_feuille]

# Complétion des données
for i in range(1,1000):
    feuille.cell(i,2).value = i**2

#Sauvegrade des modificaitons
fichier.save(path_excel)