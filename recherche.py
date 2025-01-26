import pandas as pd
from openpyxl import load_workbook
import time



path_excel = r"C:\Users\damie\Documents\GitHub\MON-3.1---Python-in-companies\Tests 5.xlsm"
nom_donnee = "Données"
nom_resultat = "Sommaire"

def derniere_ligne_remplie(sheet):
    """
    Retourne le numéro de la dernière ligne remplie dans la feuille.
    """
    derniere_ligne = 0
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value is not None:
                derniere_ligne = max(derniere_ligne, cell.row)
    return derniere_ligne

def derniere_colonne_remplie(sheet):
    """
    Retourne le numéro de la dernière colonne remplie dans la feuille.
    """
    derniere_colonne = 0
    for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column):
        for cell in col:
            if cell.value is not None:
                derniere_colonne = max(derniere_colonne, cell.column)
    return derniere_colonne

debut_ouverture = time.time()
print("Ouverture")
#Ouverture de la feuille
fichier = load_workbook(path_excel,keep_vba=True)
fin_ouverture = time.time()

print("Temps d'ouverture' : ",fin_ouverture-debut_ouverture)

for k in range(1,3) :
    print("__________________________________________________\nCas n°",k)

    compteur = 0
    somme = 0

    nom_donne_plus = nom_donnee + " " + str(k)
    feuille = fichier[nom_donne_plus]
    data = fichier[nom_resultat]

    print("Dimensions")

    temps_1 = time.time()

    derniere_ligne = derniere_ligne_remplie(feuille)
    derniere_colonne = derniere_colonne_remplie(feuille)

    temps_2 = time.time()

    print("Parcours")

    # Complétion des données
    compteur = 0
    for i in range(1,derniere_ligne):#last_row+1):
        for j in range(1,derniere_colonne):#last_column):
            if type(feuille.cell(i,j).value) != str:
                compteur += 1
                somme += feuille.cell(i,j).value

    temps_3 = time.time()

    print("Reporting")

    data.cell(11, 2 + k).value = compteur
    data.cell(12, 2 + k).value = somme
    data.cell(13, 2 + k).value = temps_2 - temps_1
    data.cell(14, 2 + k).value = temps_3 - temps_2
    data.cell(15, 2 + k).value = temps_3 - temps_1

    #Sauvegrade des modificaitons
debut_fermeture = time.time()
print("Lancement fermeture")

fichier.save(path_excel)

fin_fermeture = time.time()

print("Temps de fermeture : ",fin_fermeture - debut_fermeture)

print("FIN")
